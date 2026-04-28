import streamlit as st
import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook, load_workbook

# Page Config
st.set_page_config(
    page_title="AI Persona Quiz",
    page_icon="🤖",
    layout="centered",
    initial_sidebar_state="collapsed"
)

# Hide sidebar, menu, footer
st.markdown("""
<style>
    [data-testid="collapsedControl"] { display: none; }
    #MainMenu { visibility: hidden; }
    footer { visibility: hidden; }
    header { visibility: hidden; }

    .stApp {
        background: linear-gradient(135deg, #e8d5f5 0%, #f5e6ff 40%, #ede0fa 70%, #dcc8f0 100%);
        min-height: 100vh;
    }

    .welcome-title {
        text-align: center;
        font-size: 3rem;
        font-weight: 900;
        color: #2d1b4e;
        margin-top: 2rem;
        margin-bottom: 0.5rem;
    }

    .welcome-subtitle {
        text-align: center;
        font-size: 1.2rem;
        color: #5a3d7a;
        margin-bottom: 2rem;
    }

    .question-text {
        text-align: center;
        font-size: 1.8rem;
        font-weight: 800;
        color: #2d1b4e;
        margin-top: 1rem;
        margin-bottom: 2rem;
        line-height: 1.3;
    }

    .progress-text {
        text-align: center;
        font-size: 0.95rem;
        color: #7a5ba0;
        margin-bottom: 0.5rem;
        font-weight: 600;
    }

    /* Style ALL buttons as white cards */
    .stButton > button {
        background-color: #ffffff !important;
        color: #2d1b4e !important;
        border: 2px solid #e0d0f0 !important;
        border-radius: 14px !important;
        padding: 18px 16px !important;
        font-size: 1.05rem !important;
        font-weight: 600 !important;
        width: 100% !important;
        min-height: 70px !important;
        transition: all 0.2s ease !important;
        box-shadow: 0 2px 8px rgba(100, 60, 150, 0.08) !important;
    }

    .stButton > button:hover {
        background-color: #f3eafc !important;
        border-color: #b89ddb !important;
        transform: translateY(-2px) !important;
        box-shadow: 0 4px 16px rgba(100, 60, 150, 0.18) !important;
    }

    .stButton > button:active {
        background-color: #e4d5f5 !important;
        transform: translateY(0px) !important;
    }

    /* Special styling for navigation buttons */
    .nav-button > button {
        background-color: #7c4dba !important;
        color: white !important;
        border: none !important;
    }

    .persona-result {
        text-align: center;
        font-size: 2.2rem;
        font-weight: 800;
        padding: 1.5rem;
        border-radius: 16px;
        margin: 1rem 0;
        color: #2d1b4e;
    }

    .persona-desc {
        text-align: center;
        font-size: 1.15rem;
        padding: 0.5rem 1.5rem;
        color: #3d2060;
        line-height: 1.6;
    }

    .result-title {
        text-align: center;
        font-size: 2.5rem;
        font-weight: 900;
        color: #2d1b4e;
        margin-top: 1rem;
    }

    /* Progress bar styling */
    .stProgress > div > div > div {
        background-color: #9b59b6 !important;
    }

    .trait-chip {
        text-align: center;
        padding: 10px 8px;
        background: rgba(255,255,255,0.7);
        border-radius: 10px;
        font-size: 0.9rem;
        color: #2d1b4e;
        font-weight: 600;
    }
</style>
""", unsafe_allow_html=True)

# Persona Definitions
PERSONAS = {
    "AI Driver": {
        "emoji": "🚀",
        "color": "#2ecc71",
        "description": "You are an AI Champion! 🚀 You actively seek out AI tools, experiment boldly, and inspire others to embrace the future. You see AI as a superpower that amplifies human potential.",
        "traits": ["Early adopter", "Tech evangelist", "Innovation leader", "Always experimenting", "Loves automation"]
    },
    "Quiet Optimizer": {
        "emoji": "🧠",
        "color": "#3498db",
        "description": "You are a Silent Strategist! 🧠 You use AI smartly behind the scenes to boost your productivity. You do not need to shout about it. Your results speak for themselves.",
        "traits": ["Efficiency-focused", "Low-key power user", "Results-driven", "Practical thinker", "Works smarter"]
    },
    "Cautious Tester": {
        "emoji": "🔍",
        "color": "#f39c12",
        "description": "You are a Curious Explorer! 🔍 You see the potential of AI but want to understand it better before diving in. You ask great questions and prefer to test the waters carefully.",
        "traits": ["Thoughtful evaluator", "Asks good questions", "Open-minded", "Risk-aware", "Steady learner"]
    },
    "Skeptic": {
        "emoji": "🛡️",
        "color": "#e74c3c",
        "description": "You are a Critical Thinker! 🛡️ You value trust, accuracy, and proven methods. You push back on hype and demand that AI proves its worth before you buy in.",
        "traits": ["Values accuracy", "Questions everything", "Trust-focused", "Detail-oriented", "Prefers proven methods"]
    },
    "Unengaged": {
        "emoji": "💎",
        "color": "#9b59b6",
        "description": "You are an Untapped Opportunity! 💎 AI has not clicked for you yet, and that is okay! Once you see how it solves YOUR specific problems, you might just surprise everyone.",
        "traits": ["Waiting for the right moment", "Focused on current tasks", "Potential game-changer", "Needs a personal use case", "Fresh perspective"]
    }
}

# Quiz Questions
QUESTIONS = [
    {
        "question": "A new AI tool is introduced at work. What is your first reaction?",
        "options": [
            ("Sign me up! I want to try it right away.", "AI Driver"),
            ("I will explore it quietly on my own time.", "Quiet Optimizer"),
            ("Interesting. I will watch some demos first.", "Cautious Tester"),
            ("I will wait and see if it actually works.", "Skeptic"),
            ("Another tool? I am fine with what I have.", "Unengaged")
        ]
    },
    {
        "question": "How do you feel about AI writing your emails?",
        "options": [
            ("Love it! I already use it for drafts.", "AI Driver"),
            ("I use it sometimes. It saves me time.", "Quiet Optimizer"),
            ("I would try it, but I would heavily edit it.", "Cautious Tester"),
            ("No thanks. It will not sound like me.", "Skeptic"),
            ("I do not see why I would need that.", "Unengaged")
        ]
    },
    {
        "question": "Your team wants to automate a repetitive task with AI. You...",
        "options": [
            ("Volunteer to lead the project!", "AI Driver"),
            ("Quietly suggest the best tool for the job.", "Quiet Optimizer"),
            ("Ask for a trial period to test it first.", "Cautious Tester"),
            ("Raise concerns about accuracy and reliability.", "Skeptic"),
            ("Let the team decide. It does not affect me much.", "Unengaged")
        ]
    },
    {
        "question": "When you hear 'artificial intelligence', you think...",
        "options": [
            ("Endless possibilities and innovation!", "AI Driver"),
            ("A useful tool when applied correctly.", "Quiet Optimizer"),
            ("Promising, but I need to learn more.", "Cautious Tester"),
            ("Overhyped and potentially risky.", "Skeptic"),
            ("Not really something that affects my day-to-day.", "Unengaged")
        ]
    },
    {
        "question": "A coworker shows you a cool AI trick. You...",
        "options": [
            ("Get excited and ask them to teach you more!", "AI Driver"),
            ("Take a mental note to try it later.", "Quiet Optimizer"),
            ("Think it is neat but wonder about the limitations.", "Cautious Tester"),
            ("Wonder if it is really reliable enough to use.", "Skeptic"),
            ("Smile and nod. It is not really your thing.", "Unengaged")
        ]
    },
    {
        "question": "Your company offers free AI training. You...",
        "options": [
            ("Sign up immediately and encourage your team!", "AI Driver"),
            ("Enroll quietly. Free skill upgrade, why not?", "Quiet Optimizer"),
            ("Look at the curriculum first to see if it is relevant.", "Cautious Tester"),
            ("Pass. I would rather spend time on proven skills.", "Skeptic"),
            ("Did not notice the announcement.", "Unengaged")
        ]
    },
    {
        "question": "How frequently do you use AI tools for work-related tasks?",
        "options": [
            ("Daily. It is part of my workflow.", "AI Driver"),
            ("A few times a week for specific tasks.", "Quiet Optimizer"),
            ("Occasionally. I am still figuring them out.", "Cautious Tester"),
            ("Rarely. I do not fully trust the output.", "Skeptic"),
            ("Never. I have not felt the need.", "Unengaged")
        ]
    },
    {
        "question": "AI makes a mistake on a task. Your reaction?",
        "options": [
            ("Expected! I will tweak the prompt and try again.", "AI Driver"),
            ("Note the limitation and adjust my approach.", "Quiet Optimizer"),
            ("This is why I always double-check the output.", "Cautious Tester"),
            ("See, this is exactly why I do not trust it.", "Skeptic"),
            ("I would not know. I do not use it.", "Unengaged")
        ]
    },
    {
        "question": "If AI could do 50 percent of your job, you would feel...",
        "options": [
            ("Thrilled! More time for creative and strategic work.", "AI Driver"),
            ("Great. I would use the extra time productively.", "Quiet Optimizer"),
            ("Mixed. Excited but a little nervous.", "Cautious Tester"),
            ("Worried about job security and quality control.", "Skeptic"),
            ("Doubtful. AI cannot really do what I do.", "Unengaged")
        ]
    },
    {
        "question": "Your manager asks for your opinion on adopting AI. You say...",
        "options": [
            ("Let us go all in! Here is my proposal.", "AI Driver"),
            ("I have some ideas. Let me share them privately.", "Quiet Optimizer"),
            ("I am open to it, but let us start with a pilot.", "Cautious Tester"),
            ("We need to be very careful about the risks.", "Skeptic"),
            ("I do not have a strong opinion on this.", "Unengaged")
        ]
    },
    {
        "question": "How do you describe AI to a friend?",
        "options": [
            ("A game-changer that is transforming everything!", "AI Driver"),
            ("A handy tool, like a smart assistant.", "Quiet Optimizer"),
            ("Interesting technology, but still a work in progress.", "Cautious Tester"),
            ("Something to approach with healthy skepticism.", "Skeptic"),
            ("Tech stuff I do not really follow.", "Unengaged")
        ]
    },
    {
        "question": "In 5 years, AI in the workplace will be...",
        "options": [
            ("Everywhere, and I will be ahead of the curve!", "AI Driver"),
            ("Very useful for those who know how to leverage it.", "Quiet Optimizer"),
            ("Mainstream, but we will still need human oversight.", "Cautious Tester"),
            ("Overpromised. Reality will not match the hype.", "Skeptic"),
            ("Probably around, but I will cross that bridge later.", "Unengaged")
        ]
    }
]

# Excel Save Function
EXCEL_FILE = "ai_persona_results.xlsx"

def save_to_excel(user_name, persona, scores):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "AI Persona Results"
        headers = [
            "Timestamp", "User Name", "AI Persona Result",
            "AI Driver Score", "Quiet Optimizer Score",
            "Cautious Tester Score", "Skeptic Score", "Unengaged Score"
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)
    row = [
        timestamp, user_name, persona,
        scores.get("AI Driver", 0), scores.get("Quiet Optimizer", 0),
        scores.get("Cautious Tester", 0), scores.get("Skeptic", 0),
        scores.get("Unengaged", 0)
    ]
    ws.append(row)
    wb.save(EXCEL_FILE)

# Session State Init
if "page" not in st.session_state:
    st.session_state.page = "welcome"
if "current_q" not in st.session_state:
    st.session_state.current_q = 0
if "answers" not in st.session_state:
    st.session_state.answers = {}
if "user_name" not in st.session_state:
    st.session_state.user_name = ""
if "result_persona" not in st.session_state:
    st.session_state.result_persona = None
if "result_scores" not in st.session_state:
    st.session_state.result_scores = {}

# ─── WELCOME PAGE ──────────────────────────────────────
if st.session_state.page == "welcome":
    st.markdown("")
    st.markdown("")
    st.markdown('<div class="welcome-title">🤖 What is Your AI Persona?</div>', unsafe_allow_html=True)
    st.markdown('<div class="welcome-subtitle">Discover your AI personality at work in just 12 quick questions!</div>', unsafe_allow_html=True)

    st.markdown("")

    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("")
        name = st.text_input("👤 Enter your name to begin:", placeholder="e.g., Jade Ryan", label_visibility="visible")
        st.markdown("")

        st.markdown("**🎭 The 5 AI Personas:**")
        for pname, info in PERSONAS.items():
            st.markdown(f"{info['emoji']} **{pname}**")

        st.markdown("")
        st.markdown("")

        if st.button("🚀 Start Quiz!", use_container_width=True):
            if name.strip():
                st.session_state.user_name = name.strip()
                st.session_state.page = "quiz"
                st.session_state.current_q = 0
                st.session_state.answers = {}
                st.rerun()
            else:
                st.warning("Please enter your name to start!")

# ─── QUIZ PAGE ─────────────────────────────────────────
elif st.session_state.page == "quiz":
    q_idx = st.session_state.current_q
    q = QUESTIONS[q_idx]
    total = len(QUESTIONS)

    # Progress
    st.markdown("")
    st.markdown(f'<div class="progress-text">Question {q_idx + 1} of {total}</div>', unsafe_allow_html=True)
    st.progress((q_idx + 1) / total)

    # Question
    st.markdown(f'<div class="question-text">{q["question"]}</div>', unsafe_allow_html=True)

    # Answer buttons in 2-column grid
    options = q["options"]
    num_options = len(options)

    for row_start in range(0, num_options, 2):
        cols = st.columns(2)
        for col_idx in range(2):
            opt_idx = row_start + col_idx
            if opt_idx < num_options:
                opt_text, opt_persona = options[opt_idx]
                with cols[col_idx]:
                    if st.button(opt_text, key=f"q{q_idx}_opt{opt_idx}", use_container_width=True):
                        st.session_state.answers[q_idx] = opt_persona
                        if q_idx < total - 1:
                            st.session_state.current_q = q_idx + 1
                        else:
                            # Calculate results
                            scores = {p: 0 for p in PERSONAS}
                            for persona_val in st.session_state.answers.values():
                                scores[persona_val] += 1
                            winning_persona = max(scores, key=scores.get)
                            save_to_excel(st.session_state.user_name, winning_persona, scores)
                            st.session_state.result_persona = winning_persona
                            st.session_state.result_scores = scores
                            st.session_state.page = "result"
                        st.rerun()

    # Back button
    st.markdown("")
    st.markdown("")
    if q_idx > 0:
        back_col1, back_col2, back_col3 = st.columns([1, 1, 1])
        with back_col2:
            if st.button("⬅️ Go Back", key="back_btn", use_container_width=True):
                st.session_state.current_q = q_idx - 1
                st.rerun()

# ─── RESULT PAGE ───────────────────────────────────────
elif st.session_state.page == "result":
    persona = st.session_state.result_persona
    scores = st.session_state.result_scores
    info = PERSONAS[persona]

    st.balloons()

    st.markdown("")
    st.markdown('<div class="result-title">🎉 Your Result Is In!</div>', unsafe_allow_html=True)
    st.markdown("")

    st.markdown(
        f'<div class="persona-result" style="background: linear-gradient(135deg, {info["color"]}33, {info["color"]}55); border: 3px solid {info["color"]};">'
        f'{info["emoji"]} {persona} {info["emoji"]}'
        f'</div>',
        unsafe_allow_html=True
    )

    st.markdown(f'<div class="persona-desc">{info["description"]}</div>', unsafe_allow_html=True)

    st.markdown("")
    st.markdown("")

    # Traits
    st.markdown('<div class="progress-text">🏷️ Your Key Traits</div>', unsafe_allow_html=True)
    trait_cols = st.columns(len(info["traits"]))
    for i, trait in enumerate(info["traits"]):
        with trait_cols[i]:
            st.markdown(f'<div class="trait-chip">{trait}</div>', unsafe_allow_html=True)

    st.markdown("")
    st.markdown("")

    # Score Breakdown
    st.markdown('<div class="progress-text">📊 Score Breakdown</div>', unsafe_allow_html=True)
    score_df = pd.DataFrame({
        "Persona": list(scores.keys()),
        "Score": list(scores.values())
    }).set_index("Persona")
    st.bar_chart(score_df)

    st.success(f"Results saved for {st.session_state.user_name}!")

    st.markdown("")

    # Action buttons
    col1, col2 = st.columns(2)
    with col1:
        if st.button("🔄 Take Quiz Again", use_container_width=True):
            st.session_state.page = "welcome"
            st.session_state.current_q = 0
            st.session_state.answers = {}
            st.session_state.result_persona = None
            st.session_state.result_scores = {}
            st.session_state.user_name = ""
            st.rerun()
    with col2:
        if os.path.exists(EXCEL_FILE):
            with open(EXCEL_FILE, "rb") as f:
                st.download_button(
                    label="📥 Download Results Excel",
                    data=f,
                    file_name="ai_persona_results.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
